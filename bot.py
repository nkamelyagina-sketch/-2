import os
import re
import io
import threading
from flask import Flask
import telebot
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

# Настройка веб-сервера Flask (нужен для хостинга Render)
app = Flask('')

@app.route('/')
def home():
    return "Бот запущен и работает!"

def run_flask():
    port = int(os.environ.get("PORT", 8080))
    app.run(host='0.0.0.0', port=port)

# Инициализация бота через переменную окружения
TOKEN = os.environ.get('BOT_TOKEN')
bot = telebot.TeleBot(TOKEN)

HEADER_REGEX = r'(#\d+\s+(?:\d+\s+[а-яёА-ЯЁ]+|Вчера|Сегодня)\s+в\s+\d{1,2}:\d{2}\s+@\s*[^\n]+)'

def parse_posts(text):
    matches = list(re.finditer(HEADER_REGEX, text, re.IGNORECASE))
    if not matches:
        return []
        
    parsed_posts = []
    for idx, match in enumerate(matches):
        header = match.group(0)
        start_pos = match.end()
        end_pos = matches[idx+1].start() if idx + 1 < len(matches) else len(text)
        body = text[start_pos:end_pos].strip()
        
        body = re.sub(r'Ответить\s*\|\s*Цитировать', '', body, flags=re.IGNORECASE).strip()
        
        header_match = re.search(r'#(\d+)\s+(.*?)\s+в\s+(\d+:\d+)\s+@\s*(.*?)(?:Жалоба)?$', header, re.IGNORECASE)
        if header_match:
            post_num = int(header_match.group(1))
            date_time = f"{header_match.group(2)} в {header_match.group(3)}"
            nick = header_match.group(4).strip()
        else:
            post_num = 0
            date_time = "Неизвестно"
            nick = "Неизвестно"
            
        id_match = re.search(r'\[(и?\d+)\]', body)
        player_id = id_match.group(1) if id_match else ""
        
        body_lower = body.lower()
        team = "Иное"
        if any(w in body_lower for w in ["совята", "совёнок", "совенок"]):
            team = "Совята"
        elif any(w in body_lower for w in ["жаворонки", "жаворонок"]):
            team = "Жаворонки"
            
        links = re.findall(r'https?://[^\s]+', body)
        links_str = ", ".join(links) if links else ""
        
        activity, details, qty = analyze_activity(body_lower, body, links)
        
        parsed_posts.append({
            'num': post_num,
            'date_time': date_time,
            'nick': nick,
            'id': player_id,
            'team': team,
            'activity': activity,
            'details': details,
            'links': links_str,
            'qty': qty,
            'raw_body': body
        })
        
    return parsed_posts

def analyze_activity(body_lower, raw_body, links):
    activity = "Иное"
    details = raw_body.split('\n')[0] if raw_body else ""
    qty = 1
    
    if "патруль" in body_lower:
        activity = "Патруль"
        qty = 1
    elif "дозор" in body_lower:
        activity = "Дозор"
        if "начало дозора" in body_lower:
            qty = 0
        elif "конец дозора" in body_lower:
            qty = 1
            time_match = re.search(r'(?:время дозора|время)[^:]*:\s*([^\n]+)', body_lower)
            time_text = time_match.group(1) if time_match else body_lower
            
            hours, minutes = 0, 0
            hr_match = re.search(r'(\d+)\s*(?:час|ч)', time_text)
            if hr_match:
                hours = int(hr_match.group(1))
            min_match = re.search(r'(\d+)\s*(?:минут|м)', time_text)
            if min_match:
                minutes = int(min_match.group(1))
                
            total_mins = hours * 60 + minutes
            if total_mins > 0:
                qty = total_mins // 30
                
    elif "охота" in body_lower:
        activity = "Охота"
        items_count = 0
        keywords = ["красив", "редк", "светляч", "обычн", "упитк", "бабочк"]
        for word in keywords:
            pattern = rf'{word}[а-яё]*\s*[^0-9\n]*\s*(\d+)'
            matches = re.findall(pattern, body_lower)
            if matches:
                items_count += sum(map(int, matches))
        
        qty = items_count if items_count > 0 else max(1, len(links))
            
    elif "поручение" in body_lower:
        activity = "Поручение"
        qty = 1
    elif "ошибки" in body_lower or "поиск ошибок" in body_lower:
        activity = "Ошибки"
        qty = 1
    elif "значок" in body_lower or "значки" in body_lower:
        activity = "Значок"
        qty = 0
    elif "пополнить ряды" in body_lower or "принять участие" in body_lower:
        activity = "Вступление/Активность"
        qty = 0
        
    return activity, details, qty

def generate_excel(parsed_posts):
    wb = openpyxl.Workbook()
    
    font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_regular = Font(name="Calibri", size=11)
    font_bold = Font(name="Calibri", size=11, bold=True)
    
    fill_blue_header = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin', color='BFBFBF'),
        right=Side(style='thin', color='BFBFBF'),
        top=Side(style='thin', color='BFBFBF'),
        bottom=Side(style='thin', color='BFBFBF')
    )
    
    align_center = Alignment(horizontal='center', vertical='center')
    align_left = Alignment(horizontal='left', vertical='center')
    
    # 1. СПРАВОЧНИК
    ws_ref = wb.active
    ws_ref.title = "Справочник"
    ws_ref.views.sheetView[0].showGridLines = True
    
    ref_headers = ["Активность", "Базовый балл", "Максимум"]
    for col_num, header in enumerate(ref_headers, 1):
        cell = ws_ref.cell(row=1, column=col_num, value=header)
        cell.font = font_header
        cell.fill = fill_blue_header
        cell.alignment = align_center
        
    ref_data = [
        ["Патруль", 3, ""],
        ["Дозор", 1, ""],
        ["Охота", 1, 10],
        ["Поручение", 5, 15],
        ["Лекция", 6, ""],
        ["Ошибки", 1, ""],
        ["Иное", 0, ""]
    ]
    
    for row_num, row_data in enumerate(ref_data, 2):
        for col_num, val in enumerate(row_data, 1):
            cell = ws_ref.cell(row=row_num, column=col_num, value=val)
            cell.font = font_regular
            cell.border = thin_border
            if col_num > 1:
                cell.alignment = align_center

    # 2. БАЗА ОТЧЕТОВ
    ws_base = wb.create_sheet(title="База_отчетов")
    ws_base.views.sheetView[0].showGridLines = True
    
    base_headers = [
        "№ Отчета", "Дата/Время отписи", "Ник", "ID", 
        "Команда", "Тип активности", "Детали", "Ссылки", 
        "Кол-во / Блоки 30 мин", "Авто-Балл", "Корректировка (вручную)", "Итоговый балл"
    ]
    
    for col_num, header in enumerate(base_headers, 1):
        cell = ws_base.cell(row=1, column=col_num, value=header)
        cell.font = font_header
        cell.fill = fill_blue_header
        cell.alignment = align_center
        
    for idx, post in enumerate(parsed_posts, 2):
        ws_base.cell(row=idx, column=1, value=post['num'])
        ws_base.cell(row=idx, column=2, value=post['date_time'])
        ws_base.cell(row=idx, column=3, value=post['nick'])
        ws_base.cell(row=idx, column=4, value=post['id'])
        ws_base.cell(row=idx, column=5, value=post['team'])
        ws_base.cell(row=idx, column=6, value=post['activity'])
        ws_base.cell(row=idx, column=7, value=post['details'])
        ws_base.cell(row=idx, column=8, value=post['links'])
        ws_base.cell(row=idx, column=9, value=post['qty'])
        
        ws_base.cell(row=idx, column=10, value=(
            f'=IF(F{idx}="Патруль", Справочник!$B$2, '
            f'IF(F{idx}="Дозор", I{idx}*Справочник!$B$3, '
            f'IF(F{idx}="Охота", MIN(Справочник!$C$4, I{idx}*Справочник!$B$4), '
            f'IF(F{idx}="Поручение", MIN(Справочник!$C$5, I{idx}*Справочник!$B$5), '
            f'IF(F{idx}="Лекция", I{idx}*Справочник!$B$6, '
            f'IF(F{idx}="Ошибки", I{idx}*Справочник!$B$7, I{idx}))))))'
        ))
        
        ws_base.cell(row=idx, column=12, value=f'=IF(ISBLANK(K{idx}), J{idx}, K{idx})')

    for r in range(2, len(parsed_posts) + 2):
        for c in range(1, 13):
            cell = ws_base.cell(row=r, column=c)
            cell.font = font_regular
            cell.border = thin_border
            if c in [1, 4, 5, 6, 9, 10, 11, 12]:
                cell.alignment = align_center
            else:
                cell.alignment = align_left

    # 3. ИТОГОВАЯ ТАБЛИЦА
    ws_total = wb.create_sheet(title="Итоговая_таблица")
    ws_total.views.sheetView[0].showGridLines = True
    
    total_headers = [
        "Ник игрока", "ID", "Команда", 
        "Баллы Патруль", "Баллы Дозор", "Баллы Охота", 
        "Баллы Поручение", "Баллы Ошибки", "Баллы Иное", "ВСЕГО БАЛЛОВ"
    ]
    
    for col_num, header in enumerate(total_headers, 1):
        cell = ws_total.cell(row=1, column=col_num, value=header)
        cell.font = font_header
        cell.fill = fill_blue_header
        cell.alignment = align_center

    players = {}
    for post in parsed_posts:
        nick = post['nick']
        if nick == "Неизвестно" or not nick:
            continue
        if nick not in players:
            players[nick] = {'id': post['id'], 'team': post['team']}
        if post['id'] and not players[nick]['id']:
            players[nick]['id'] = post['id']
        if post['team'] != 'Иное' and players[nick]['team'] == 'Иное':
            players[nick]['team'] = post['team']

    for r_idx, (nick, info) in enumerate(players.items(), 2):
        ws_total.cell(row=r_idx, column=1, value=nick)
        ws_total.cell(row=r_idx, column=2, value=info['id'])
        ws_total.cell(row=r_idx, column=3, value=info['team'])
        
        max_row = len(parsed_posts) + 2
        ws_total.cell(row=r_idx, column=4, value=f'=SUMIFS(База_отчетов!$L$2:$L${max_row}, База_отчетов!$C$2:$C${max_row}, A{r_idx}, База_отчетов!$F$2:$F${max_row}, "Патруль")')
        ws_total.cell(row=r_idx, column=5, value=f'=SUMIFS(База_отчетов!$L$2:$L${max_row}, База_отчетов!$C$2:$C${max_row}, A{r_idx}, База_отчетов!$F$2:$F${max_row}, "Дозор")')
        ws_total.cell(row=r_idx, column=6, value=f'=SUMIFS(База_отчетов!$L$2:$L${max_row}, База_отчетов!$C$2:$C${max_row}, A{r_idx}, База_отчетов!$F$2:$F${max_row}, "Охота")')
        ws_total.cell(row=r_idx, column=7, value=f'=SUMIFS(База_отчетов!$L$2:$L${max_row}, База_отчетов!$C$2:$C${max_row}, A{r_idx}, База_отчетов!$F$2:$F${max_row}, "Поручение")')
        ws_total.cell(row=r_idx, column=8, value=f'=SUMIFS(База_отчетов!$L$2:$L${max_row}, База_отчетов!$C$2:$C${max_row}, A{r_idx}, База_отчетов!$F$2:$F${max_row}, "Ошибки")')
        ws_total.cell(row=r_idx, column=9, value=f'=SUMIFS(База_отчетов!$L$2:$L${max_row}, База_отчетов!$C$2:$C${max_row}, A{r_idx}, База_отчетов!$F$2:$F${max_row}, "Иное")')
        ws_total.cell(row=r_idx, column=10, value=f'=SUM(D{r_idx}:I{r_idx})')

    for r in range(2, len(players) + 2):
        for c in range(1, 11):
            cell = ws_total.cell(row=r, column=c)
            cell.font = font_regular if c < 10 else font_bold
            cell.border = thin_border
            if c > 1:
                cell.alignment = align_center
            else:
                cell.alignment = align_left

    for ws in [ws_ref, ws_base, ws_total]:
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 11)

    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    return file_stream

@bot.message_handler(commands=['start', 'help'])
def send_welcome(message):
    welcome_text = (
        "👋 Привет! Я готов принимать отчеты.\n\n"
        "📥 Просто скопируй текст отчетов с форума и пришли мне сообщением (или в .txt файле).\n"
        "📊 Я вышлю тебе готовый Excel с формулами расчета баллов!"
    )
    bot.reply_to(message, welcome_text)

@bot.message_handler(content_types=['text', 'document'])
def handle_reports(message):
    text_data = ""
    if message.content_type == 'document':
        if message.document.file_name.endswith('.txt'):
            file_info = bot.get_file(message.document.file_id)
            downloaded_file = bot.download_file(file_info.file_path)
            text_data = downloaded_file.decode('utf-8')
        else:
            bot.reply_to(message, "⚠️ Пришли файл .txt или обычное текстовое сообщение.")
            return
    else:
        text_data = message.text

    parsed_data = parse_posts(text_data)
    if not parsed_data:
        bot.reply_to(message, "❌ Не удалось распознать формат. Отправьте сообщения с заголовками типа `#1 9 июля...`")
        return

    msg = bot.send_message(message.chat.id, f"⚙️ Распознано {len(parsed_data)} отчетов. Генерирую Excel файл...")
    
    try:
        excel_file = generate_excel(parsed_data)
        excel_file.name = "Подсчет_отчетов.xlsx"
        bot.send_document(message.chat.id, excel_file, caption="📊 Готово! Все формулы настроены.")
    except Exception as e:
        bot.edit_message_text(f"❌ Ошибка генерации: {str(e)}", message.chat.id, msg.message_id)

if __name__ == '__main__':
    # Запускаем Flask в фоновом потоке
    threading.Thread(target=run_flask, daemon=True).start()
    
    # Запускаем Telegram-бота в основном потоке
    print("Бот успешно запущен!")
    bot.infinity_polling()